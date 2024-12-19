import json
from typing import List
import openpyxl
from openpyxl.drawing.image import Image
import cv2
from collections import defaultdict
import pandas as pd
import matplotlib.pyplot as plt
import os
import math


def read_json(rpath: str):
    with open(rpath, 'r', encoding='utf-8') as file:
        data = json.load(file)
        return data


def write_json(anns: List, wpath: str):
    json.dump(anns, open(wpath, "w"))


def add_visilize2screenshot(image_rpath, ann):
    if ann["action_type_text"] == "click":
        touch_point, lift_point = ann["touch"], ann["lift"]
        click_point = [(touch_point[0] + lift_point[0]) / 2, (touch_point[1] + lift_point[1]) / 2]

        image = cv2.imread(image_rpath)
        height, width, _ = image.shape

        x = int(click_point[0] * width)
        y = int(click_point[1] * height)

        cv2.circle(image, (x, y), 20, (0, 0, 255), -1)

        image_wpath = image_rpath.split(".")[0] + "_modify.png"
        cv2.imwrite(image_wpath, image) 
        return image_wpath
    else:
        return image_rpath


def write_to_excel(anns, wpath):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="Image")
    ws.cell(row=1, column=2, value="Task")
    ws.cell(row=1, column=3, value="Action")
    ws.cell(row=1, column=4, value="Response")
    ws.cell(row=1, column=5, value="Rating")

    for idx, ann in enumerate(anns, start=2):
        ws.cell(row=idx, column=2, value=ann["task"])
        ws.cell(row=idx, column=3, value=ann["action"])
        ws.cell(row=idx, column=4, value=ann["response"])
        ws.cell(row=idx, column=5, value=ann["rating"])

        img = Image(ann["image_path"])
        img.width, img.height = (240, 480)
        ws.row_dimensions[idx].height = 400
        ws.add_image(img, f'A{idx}')

    ws.column_dimensions['A'].width = 20
    wb.save(wpath)


def parse_rating(ann):
    try:
        response = ann["response"].replace("*", "")
        rating = int(response.split("Rating: ")[1].split("\n")[0])
        return rating
    except:
        return -1
    

def write_jsonl(anns, wpath):
    with open(wpath, 'w', encoding='utf - 8') as f:
        for item in anns:
            json_line = json.dumps(item)
            f.write(json_line + '\n')


def read_jsonl(rpath):
    data = []
    with open(rpath, 'r', encoding='utf - 8') as f:
        for idx, line in enumerate(f):
            line = line.strip()
            try:
                json_obj = json.loads(line)
                data.append(json_obj)
            except:
                print(f"Error decoding JSON on line: {idx}")
    return data


def read_xlsx(rpath):
    data = pd.read_excel(rpath)
    return data.to_dict(orient="records")


def sample_data(rpath, wpath):
    sample_anns = []
    statistics = defaultdict(int)
    last_step_statistics = defaultdict(int)
    anns = read_jsonl(rpath)
    for ann in anns:
        statistics[ann["rating"]] += 1
        if ann["step_id"] == len(ann["action_list"]) - 1:
            last_step_statistics[ann["rating"]] += 1
            sample_anns.append(ann)

    print(statistics)
    print(last_step_statistics)

    # sample_anns = random.sample(anns, 100)
    write_to_excel(sample_anns, wpath)


def dict_mean(dict_list):
    mean_dict = {}
    if len(dict_list) > 0:
        for key in dict_list[0].keys():
            if "min" in key:
                mean_dict[key] = min(d[key] for d in dict_list)
            elif "max" in key:
                mean_dict[key] = max(d[key] for d in dict_list)
            else:
                mean_dict[key] = sum(d[key] for d in dict_list) / len(dict_list)
    return mean_dict


def smooth(scalars: List[float]) -> List[float]:
    if len(scalars) == 0:
        return []

    last = scalars[0]
    smoothed = []
    weight = 1.8 * (1 / (1 + math.exp(-0.05 * len(scalars))) - 0.5)  # a sigmoid function
    for next_val in scalars:
        smoothed_val = last * weight + (1 - weight) * next_val
        smoothed.append(smoothed_val)
        last = smoothed_val
    return smoothed


def plot_loss(log_dir: str, keys: List[str] = ["loss"]) -> None:
    plt.switch_backend("agg")
    data = read_jsonl(os.path.join(log_dir, "train_log.jsonl"))

    for key in keys:
        steps, metrics = [], []
        for i in range(len(data)):
            if key in data[i]:
                steps.append(data[i]["step"])
                metrics.append(data[i][key])

        plt.figure()
        plt.plot(steps, metrics, color="#1f77b4", alpha=0.4, label="original")
        plt.plot(steps, smooth(metrics), color="#1f77b4", label="smoothed")
        plt.title(f"{key} of {log_dir}")
        plt.xlabel("step")
        plt.ylabel(key)
        plt.legend()
        figure_path = os.path.join(log_dir, "training_{}.png".format(key))
        plt.savefig(figure_path, format="png", dpi=100)
        print("Figure saved at:", figure_path)


# print("-------")
# anns = read_jsonl("data/aitw_anns/1218/aitw_val_score.jsonl")
# anns_2 = []
# for ann in anns:
#     if ann["rating"] == 2:
#         anns_2.append(ann)
# write_to_excel(anns=anns_2[:20], wpath="val_score_2.xlsx")