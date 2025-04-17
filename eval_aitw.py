import argparse
import yaml
import json
import os
import cv2
import openpyxl
from openpyxl.drawing.image import Image
import cv2

import utils
from dataset import create_dataset
from models import create_agent
from data_preprocess.utils import ActionType, to_autoui
from eval_tools.androidenv import AndroidEnv


# 在截图上添加可视化的触摸点
def add_visilize2screenshot(image_rpath, action):
    # 如果动作类型不是双点触控，直接返回原路径
    if action.action_type != ActionType.DualPoint:
        return image_rpath

    # 读取图片并获取其尺寸
    image = cv2.imread(image_rpath)
    height, width, _ = image.shape

    # 根据触摸点的比例计算实际像素位置
    x = int(action.touch_point[0] * width)
    y = int(action.touch_point[1] * height)

    # 在图片上绘制一个红色圆点表示触摸点
    cv2.circle(image, (x, y), 50, (0, 0, 255), -1)

    # 保存修改后的图片并返回新路径
    image_wpath = image_rpath.replace(".png", "") + f"_point.png"
    cv2.imwrite(image_wpath, image) 

    return image_wpath


# 评估函数，执行任务并记录结果
def evaluation(config, agent, dataset, env, ann_wpath):
    # 打开结果文件以追加模式写入
    with open(ann_wpath, "a") as fout:
        # 遍历数据集中的任务
        for task_id, task, query_format in dataset:
            done, history = False, []  # 初始化任务完成状态和历史记录

            step_num = 0
            # 获取初始截图路径
            current_screenshot_path = env.get_obs(task_id, step_num)
            while not done:
                step_num += 1
                # 根据模型类型生成提示文本
                if config["model_name"] == "cogagent":
                    text = query_format.format(task, "".join(history))
                else:
                    text = query_format.format("\n".join(history), task)
                
                # 获取智能体的动作
                raw_action = agent.get_action(text=text, image_path=current_screenshot_path)
                # 翻译动作为环境可识别的格式
                action = env.translate_action(raw_action)
                # 在截图上添加动作点的可视化
                point_image_path = add_visilize2screenshot(current_screenshot_path, action)

                # 执行动作并获取下一步的截图路径、完成状态等信息
                next_screenshot_path, done, action, explanation = env.step(task_id, step_num, task, raw_action)
                
                # 根据模型类型更新历史记录
                if config["model_name"] == "cogagent":
                    pass
                    # history.append(f"\n{step_num-1}. {grounded_operation}\t{action_description}")
                elif config["model_name"] == "autogui":
                    action_desc = to_autoui(action, all_dict=True)
                    history.append(action_desc)
                else:
                    raise KeyError
                
                # 打印当前步骤的信息
                print("============")
                print(f"{task_id}: {task}")
                print(f"current_image: {current_screenshot_path}")
                print(f"action: {action}")
                print(f"point action on image: {point_image_path}")
                print(f"next screen shot: {next_screenshot_path}")
                print(f"if done: {done}")
                
                # 构造当前步骤的结果字典
                result ={
                    "task_id": task_id,
                    "step_id": step_num,
                    "task": task,
                    "action": history[-1],
                    "current_image_path": current_screenshot_path.replace("\\", "/"),
                    "point_image_path": point_image_path.replace("\\", "/"),
                    "next_image_path": next_screenshot_path.replace("\\", "/"),
                    "if_done": done,
                    "prompt": text,
                    "gpt-4o": explanation
                }

                # 将当前步骤的结果以JSON格式写入文件
                fout.writelines(json.dumps(result) + "\n")

                # 更新当前截图路径为下一步的截图路径
                current_screenshot_path = next_screenshot_path

                # 如果步骤数超过10，或任务完成，或动作类型为TaskComplete，则终止当前任务
                if step_num > 10 or done or action.action_type == ActionType.TaskComplete:
                    # 按下Home键（keycode 3）回到主屏幕
                    env.driver.press_keycode(3)
                    break
                
    # 关闭输出文件
    fout.close()


# 主函数，负责初始化配置、数据集、环境和智能体，并调用评估函数
def main(config):
    # 打印配置信息
    print("config:", json.dumps(config))
    
    # 根据配置中的output_name生成输出文件路径
    ann_wpath = f"./data/{config['output_name']}.jsonl"
    
    # 初始化完成任务信息的字典、成功任务数和步骤长度
    finish_task, success_num, step_len = {}, 0, 0
    
    # 如果输出文件已存在，读取并处理已有的评估结果
    if os.path.exists(ann_wpath):
        # 遍历读取的JSONL文件中的每条记录
        for ann in utils.read_jsonl(ann_wpath):
            # 如果当前任务不在完成任务字典中，则初始化
            if ann["task"] not in finish_task:
                finish_task[ann["task"]] = {"success": False, "steps": []}
            # 如果任务标记为完成，更新success状态
            if ann["if_done"]:
                finish_task[ann["task"]]["success"] = True
            # 将当前步骤信息添加到对应任务的steps列表
            finish_task[ann["task"]]["steps"].append(ann)

        # 统计成功任务数和总步骤数
        for _, info in finish_task.items():
            if info["success"]: success_num += 1
            step_len += len(info["steps"])

        # 将统计结果写入JSON文件
        utils.write_json({"success_num": success_num, "step_num": step_len, "info": finish_task}, ann_wpath.replace("jsonl", "json"))
        
        # 如果有完成的任务，打印统计信息
        if len(finish_task.keys()) > 0:
            print(f"### finish task num: {len(finish_task.keys())}\tsuccess: {success_num}\tstep_len: {step_len/len(finish_task.keys())}")

    # 打印输出文件路径
    print("output_path: ", ann_wpath)

    # 创建数据集
    print("Creating datasets")
    dataset = create_dataset(config, finish_task)

    # 构建Android环境
    print("build android env")
    env = AndroidEnv(config)

    # 创建智能体
    print("Creating agent")
    agent = create_agent(config)

    print("### Start evaluating")

    # 调用评估函数
    evaluation(config, agent, dataset, env, ann_wpath)


# 将评估结果写入Excel文件
def write_to_excel(anns, wpath):
    wb = openpyxl.Workbook()
    ws = wb.active

    # 设置表头
    ws.cell(row=1, column=1, value="current image")
    ws.cell(row=1, column=2, value="current image(add point)")
    ws.cell(row=1, column=3, value="next image")
    ws.cell(row=1, column=4, value="task")
    ws.cell(row=1, column=5, value="action")
    ws.cell(row=1, column=6, value="prompt")
    ws.cell(row=1, column=7, value="if_done")
    ws.cell(row=1, column=8, value="explanation")

    # 遍历评估结果并写入Excel
    for idx, ann in enumerate(anns, start=2):
        ws.cell(row=idx, column=4, value=ann["task"])
        ws.cell(row=idx, column=5, value=ann["action"])
        ws.cell(row=idx, column=6, value=ann["prompt"])
        ws.cell(row=idx, column=7, value=ann["if_done"])
        ws.cell(row=idx, column=8, value=ann["gpt-4o"])
        
        # 添加图片到Excel
        img = Image(ann["current_image_path"].replace("\\", "/"))
        img.width, img.height = (240, 480)
        ws.row_dimensions[idx].height = 400
        ws.add_image(img, f'A{idx}')
        img = Image(ann["point_image_path"])
        img.width, img.height = (240, 480)
        ws.row_dimensions[idx].height = 400
        ws.add_image(img, f'B{idx}')
        img = Image(ann["next_image_path"])
        img.width, img.height = (240, 480)
        ws.row_dimensions[idx].height = 400
        ws.add_image(img, f'C{idx}')


    # 设置列宽并保存Excel文件
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    wb.save(wpath)


# 程序入口
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('--model', type=str, required=True)

    args = parser.parse_args()

    # 加载配置文件
    config = f"configs/android_eval/online_eval_{args.model}.yaml"
    with open(config, 'r') as file:
        config = yaml.safe_load(file)

    # 调用主函数
    main(config)

    # 示例代码：将评估结果写入Excel
    # anns = utils.read_jsonl("data/digirl_webshop_offline.jsonl")
    # write_to_excel(anns[:100], "our.xlsx")
